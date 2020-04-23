import requests
import config
import PyPDF2
import openpyxl
import numpy as np
import ibm_db


# steps
# pip install pipenv
# pipenv install requests
# zatvori/otori vs code, ako klikom dole lijevo se ne nudi env verzija pythona od projekta
# promijenis verziju pythona da koristi env python
# vs code ponudi da instaliras pylint, kliknes install
# ctrl+alt+f formatiraj text--> vs code install autopep8

# radi sve ok, ali se javlja greska priliko installa pylint i autopep8...
# greska, rijesi da se ne pojvaljuje
# virtualenv Scripts\Activate.ps1 cannot be loaded because running scripts is disabled
# rjesenje https://github.com/microsoft/vscode-python/issues/2559
# u cmd: Set-ExecutionPolicy -ExecutionPolicy RemoteSigned -Scope CurrentUser
# u cmd: Get-ExecutionPolicy -LIST
# i onda ti se otvori python sa prefixom (naziv env projekta)

url = "https://api.github.com/users/test"
headers = {
    "Authorization": "Bearer " + config.api_key
}
params = {
    "dummy": "test"
}
response = requests.get(url)  # get(url, headers=headers, params=params)
print(response.json()["avatar_url"])  # extract value from dict
# reminder -> [item for item in items if item[1] > 9]

# twilio send sms
# pipenv install twilio
# skip

# web scraping, extractiong from html, xml,
# pipenv install beautiflsoup4
# response = request.get("...")
# soup = BeautifulSoup(response.txt, "html.parser")
# questions = soup.select(".question-summary") -> filter po css klasi
# print(questions[0].attrs) -> dict.. id class
# print(questions[0].get("id", 0))
# print(questions[0].select(".asdasd"))
# skip

# browser automation - tests
# pipenv selunuim
# driver za automatizaciju odredjenog browsera
# odes na stranicu od selenium paketa i vidis koji se driveri nude za chrome, edge, firefox, safari..
# skip

# pdf files
# pipenv install pypdf2
with open("first.pdf", "rb") as file: # rb -read binary
    reader = PyPDF2.PdfFileReader(file)
    print(reader.numPages)
    page = reader.getPage(0)
    page.rotateClockwise(90)
    writer = PyPDF2.PdfFileWriter()
    writer.addPage(page)
    # writer.insertPage(page, 2)
    with open("first-rotated.pdf", "wb") as output:
        writer.write(output)

merger = PyPDF2.PdfFileMerger()
file_names = ["first.pdf", "second.pdf"]
for file_name in file_names:
    merger.append(file_name)
merger.write("combined.pdf")

# excel
# wb = openpyxl.workbook()

# wb = openpyxl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)
# sheet = wb["Sheet1"]
# wb.create_sheet("Sheet2", 0) # 0 - index where to place sheet
# wb.remove_sheet()

# prvi nacin
# cell = sheet["a:1"]
# print(cell.value)
# cell.value = "New value"

# print(cell.row) # 1
# print(cell.column) # A
# print(cell.coordinate) # A1

# drugi nacin
# cell = sheet.cell(row=1, column=1) # koristi se kad hoces u for loop-u dinamicki pristupiti cell-u
# print(sheet.max_row) # 4
# print(sheet.max_column) # 3
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# column = sheet["a"] # all cells in a column, tuple
# cells = sheet["a:c"] # tuple of tuples columns
# sheet[1] # all cells in row 1
# sheet[1:3] # all cells in row 1-3

# sheet.append([1,2,3])
# insert row, column, delete.. sort, filter

# wb.save("transactions2.xlsx")

# command query separation principle
# metoda mora ili da napravi nesto i promijeni neki podatak -- command
# ili da ne mijenja nista i samo vrati neki rezultati -- query
# ne jedno i drugo

# numPy
array = np.array([1, 2, 3])
print(array) # [1 2 3]
print(type(array)) # <class 'numpy.ndarray'>

matrix_array = np.array([[1, 2, 3], [4, 5, 6]])
print(matrix_array) # [[1 2 3]
#                      [4 5 6]]
print(matrix_array.shape) # (2, 3)

array = np.zeros((3, 4), dtype=int)
# array = np.ones((3, 4), dtype=int) # 1 umjesto 0
# array = np.full((3, 4), 5, dtype=int) # 5 umjesto 0
print(array) #
# [[0 0 0 0]
#  [0 0 0 0]
#  [0 0 0 0]]

array = np.random.random((3, 4)) # random 0-1 
print(array)
# [[0.43472529 0.38865677 0.40858233 0.48367123]
#  [0.84865588 0.76132091 0.01474215 0.92611911]
#  [0.30907744 0.12683381 0.11345782 0.09641577]]
print(array[0, 0]) # 0.4347252912847549, da nije np-a, pisao bi array[0][0]
print(array > 0.2)
# [[ True  True  True  True]
#  [ True  True False  True]
#  [ True False False False]]

print(array[array > 0.2]) # returns array sa vrijednostima vecicm od 0.2
print(np.sum(array))
print(np.floor(array))
print(np.ceil(array))
print(np.round(array))

first = np.array([1, 2, 3])
second = np.array([1, 2, 3])
print(first + second) # [2 4 6]

dimensions_inch = np.array([1, 2, 3])
dimensions_cm = dimensions_inch * 2.54
print(dimensions_cm) # [2.54 5.08 7.62]

# u cistom pythonu bi to napravio
print([x * 2.54 for x in dimensions_inch])